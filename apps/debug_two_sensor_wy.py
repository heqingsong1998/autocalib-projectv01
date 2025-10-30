# -*- coding: utf-8 -*-
"""
PyQt 版多传感器监视与清零  +  原始/对齐数据落盘
- 持续读取：标准六轴(m8128b1) + 实验室六轴
- 界面显示最新值 + 滚动日志
- 按钮：清零、采集、保存、标定采样、导出、退出
- 实时曲线显示
"""

import os
import sys
import yaml
import time
import csv
import threading
from queue import Queue, Empty
from datetime import datetime
from collections import deque
from dataclasses import dataclass
from typing import Dict, Any, List, Optional

# Excel
from openpyxl import Workbook, load_workbook

# PyQt5
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QCheckBox, 
                             QGroupBox, QTextEdit, QFrame, QSizePolicy)
from PyQt5.QtCore import QTimer, Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont

# PyQtGraph
import pyqtgraph as pg

# 将项目根目录加入路径
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

# 分别为两个传感器导入各自的驱动
# 标准六轴驱动
from drivers.sensors.utils import create_sensor as create_std_sensor, initialize_sensor as init_std_sensor
# 实验室六轴驱动
from drivers.liuzhouwy.utils import create_sensor as create_lab_sensor, initialize_sensor as init_lab_sensor


def load_config():
    """加载配置文件"""
    cfg_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config", "default.yaml")
    with open(cfg_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


@dataclass
class Sample:
    ts_mono: float      # monotonic seconds
    frame: int
    data: tuple         # 原样保存传感器返回的 tuple


class RingBuffer:
    """环形缓冲器，用于存储最近的传感器数据"""
    def __init__(self, max_seconds: float = 120.0):
        self.max_seconds = max_seconds
        self._buf: Dict[str, deque] = {}

    def push(self, device: str, ts_mono: float, frame: int, data_tuple: tuple):
        dq = self._buf.setdefault(device, deque())
        dq.append(Sample(ts_mono, frame, data_tuple))
        cutoff = ts_mono - self.max_seconds
        while dq and dq[0].ts_mono < cutoff:
            dq.popleft()

    def slice(self, device: str, t_start: float, t_end: float) -> List[Sample]:
        dq = self._buf.get(device)
        if not dq:
            return []
        return [s for s in dq if (t_start <= s.ts_mono <= t_end)]

    @staticmethod
    def now_mono() -> float:
        return time.monotonic()


class RawWriter:
    """原始数据写入器"""
    def __init__(self, xlsx_path: str, flush_every_n: int = 200):
        self.xlsx_path = xlsx_path
        self.csv_liuzhou = os.path.splitext(xlsx_path)[0] + "_liuzhou_raw.csv"
        self.csv_lab_six = os.path.splitext(xlsx_path)[0] + "_lab_six_raw.csv"
        self.csv_aligned = os.path.splitext(xlsx_path)[0] + "_aligned.csv"
        self.flush_every_n = flush_every_n
        self._xlsx_lock = threading.Lock()

        # 准备 Excel
        if not os.path.exists(self.xlsx_path):
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Sheet1_Standard_Six"
            wb.create_sheet("Sheet2_Lab_Six")
            wb.save(self.xlsx_path)

        # CSV 句柄
        self._csv_files = {
            "liuzhou": open(self.csv_liuzhou, "a", newline="", encoding="utf-8"),
            "lab_six": open(self.csv_lab_six, "a", newline="", encoding="utf-8"),
            "aligned": open(self.csv_aligned, "a", newline="", encoding="utf-8"),
        }
        self._csv_writers = {
            "liuzhou": None,
            "lab_six": None,
            "aligned": None,
        }
        self._csv_headers = {
            "liuzhou": ["ts_iso", "ts_epoch", "frame", "fx", "fy", "fz", "mx", "my", "mz", "temp"],
            "lab_six": ["ts_iso", "ts_epoch", "frame", "fx", "fy", "fz", "mx", "my", "mz", "temp"],
            "aligned": [
                "ts_iso", "ts_epoch", "ts_mono", "std_frame", "lab_frame",
                "std_fx", "lab_fx",
                "std_fy", "lab_fy",
                "std_fz", "lab_fz",
                "std_mx", "lab_mx",
                "std_my", "lab_my",
                "std_mz", "lab_mz",
            ],
        }
        self._csv_counts = {
            "liuzhou": 0,
            "lab_six": 0,
            "aligned": 0,
        }
        self._csv_lock = threading.Lock()

    def append_raw_row(self, device: str, row: Dict[str, Any]):
        """将一行数据写入对应的 CSV 文件"""
        with self._csv_lock:
            if self._csv_writers[device] is None:
                self._csv_writers[device] = csv.DictWriter(
                    self._csv_files[device], fieldnames=self._csv_headers[device]
                )
                self._csv_writers[device].writeheader()
            
            for k in self._csv_headers[device]:
                if k not in row:
                    row[k] = ""
            self._csv_writers[device].writerow(row)
            self._csv_counts[device] += 1
            if (self._csv_counts[device] % self.flush_every_n) == 0:
                self._csv_files[device].flush()

    def append_aligned_row(self, row: Dict[str, Any]):
        """写入一行对齐数据到 aligned.csv"""
        device = "aligned"
        with self._csv_lock:
            if self._csv_writers[device] is None:
                self._csv_writers[device] = csv.DictWriter(
                    self._csv_files[device], fieldnames=self._csv_headers[device]
                )
                self._csv_writers[device].writeheader()

            for k in self._csv_headers[device]:
                if k not in row:
                    row[k] = ""
            self._csv_writers[device].writerow(row)
            self._csv_counts[device] += 1
            # 对齐数据用于实时查看，强制每次写入都刷新到磁盘
            self._csv_files[device].flush()

    def append_sheet2_row(self, header: List[str], values: List[Any]):
        """将标定采样的均值数据追加到 Excel 的 Sheet2_Aligned"""
        with self._xlsx_lock:
            wb = load_workbook(self.xlsx_path)
            if "Sheet2_Aligned" not in wb.sheetnames:
                ws = wb.create_sheet("Sheet2_Aligned")
                ws.append(header)
            else:
                ws = wb["Sheet2_Aligned"]
            ws.append(values)
            wb.save(self.xlsx_path)

    def export_csv_to_sheets(self):
        """将 CSV 全量导入到 Excel"""
        wb = load_workbook(self.xlsx_path)

        # 导入标准六轴数据
        if os.path.exists(self.csv_liuzhou):
            if "Sheet1_Standard_Six" in wb.sheetnames:
                wb.remove(wb["Sheet1_Standard_Six"])
            ws = wb.create_sheet("Sheet1_Standard_Six")
            with open(self.csv_liuzhou, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)

        # 导入实验室六轴数据
        if os.path.exists(self.csv_lab_six):
            if "Sheet2_Lab_Six" in wb.sheetnames:
                wb.remove(wb["Sheet2_Lab_Six"])
            ws = wb.create_sheet("Sheet2_Lab_Six")
            with open(self.csv_lab_six, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)

        # 导入对齐数据
        if os.path.exists(self.csv_aligned):
            if "Sheet3_Aligned" in wb.sheetnames:
                wb.remove(wb["Sheet3_Aligned"])
            ws = wb.create_sheet("Sheet3_Aligned")
            with open(self.csv_aligned, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)

        wb.save(self.xlsx_path)

    def close(self):
        """关闭 CSV 文件"""
        try:
            with self._csv_lock:
                for f in self._csv_files.values():
                    f.flush()
                    f.close()
        except Exception:
            pass


class PeriodicAligner:
    """基于双缓冲固定长度队列 + 周期取最新值
    - 维护两个固定长度的队列（默认 200）用于两个六轴传感器
    - 当两队列均达到最小就绪长度后（默认 200），开始以 period_ms 周期取两个队列的最新值，并写入 aligned.csv
    - 仅在 is_collecting() 为 True 时写入
    """

    def __init__(self, raw_writer: RawWriter, period_ms: int = 10, maxlen: int = 500, min_ready: int = 500,
                 is_collecting_fn=lambda: True, require_both_changed: bool = False, debug_print: bool = False):
        self.raw_writer = raw_writer
        self.period_s = max(1, int(period_ms)) / 1000.0
        self.maxlen = max(10, int(maxlen))
        self.min_ready = max(1, int(min_ready))
        self.is_collecting_fn = is_collecting_fn
        self.require_both_changed = bool(require_both_changed)
        self.debug_print = bool(debug_print)

        self._buf = {
            "liuzhou": deque(maxlen=self.maxlen),
            "lab_six": deque(maxlen=self.maxlen),
        }
        self._lock = threading.Lock()
        self._stop = threading.Event()
        self._thread = None
        self._last_std_frame = None
        self._last_lab_frame = None

    def start(self):
        if self._thread and self._thread.is_alive():
            return
        self._stop.clear()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def stop(self):
        self._stop.set()
        if self._thread:
            try:
                self._thread.join(timeout=0.5)
            except Exception:
                pass

    def add(self, name: str, ts_mono: float, frame: int, data_tuple: tuple):
        if name not in self._buf:
            return
        with self._lock:
            dq = self._buf[name]
            dq.append((ts_mono, frame, data_tuple))

    def _run(self):
        # 对齐节拍
        next_t = time.monotonic()
        while not self._stop.is_set():
            now = time.monotonic()
            if now < next_t:
                time.sleep(min(0.002, next_t - now))
                continue
            t = next_t
            next_t += self.period_s

            # 仅在采集中写入
            try:
                if not self.is_collecting_fn():
                    continue
            except Exception:
                pass

            with self._lock:
                dqA = self._buf["liuzhou"]
                dqB = self._buf["lab_six"]
                if (len(dqA) < self.min_ready) or (len(dqB) < self.min_ready):
                    continue
                t0, fA, vA = dqA[-1]
                t1, fB, vB = dqB[-1]

            # 帧变化门控：减少周期采样导致的重复
            if self.require_both_changed:
                if (self._last_std_frame is not None) and (self._last_lab_frame is not None):
                    if (fA == self._last_std_frame) or (fB == self._last_lab_frame):
                        continue
            else:
                if (self._last_std_frame is not None) and (self._last_lab_frame is not None):
                    if (fA == self._last_std_frame) and (fB == self._last_lab_frame):
                        continue

            # 两个传感器都至少有6个通道 (fx..mz)
            if len(vA) < 6 or len(vB) < 6:
                continue
            
            std_fx, std_fy, std_fz, std_mx, std_my, std_mz = vA[:6]
            lab_fx, lab_fy, lab_fz, lab_mx, lab_my, lab_mz = vB[:6]

            ts_iso = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            ts_epoch = f"{time.time():.6f}"

            row = {
                "ts_iso": ts_iso,
                "ts_epoch": ts_epoch,
                "ts_mono": f"{t:.6f}",
                "std_frame": fA,
                "lab_frame": fB,
                "std_fx": std_fx, "lab_fx": lab_fx,
                "std_fy": std_fy, "lab_fy": lab_fy,
                "std_fz": std_fz, "lab_fz": lab_fz,
                "std_mx": std_mx, "lab_mx": lab_mx,
                "std_my": std_my, "lab_my": lab_my,
                "std_mz": std_mz, "lab_mz": lab_mz,
            }
            if self.debug_print:
                try:
                    std_vals = tuple(round(float(x), 3) for x in vA[:6])
                    lab_vals = tuple(round(float(x), 3) for x in vB[:6])
                    print(f"[ALIGN DEBUG] t={t:.3f} std#={fA} std={std_vals} | lab#={fB} lab={lab_vals}")
                except Exception:
                    pass
            self.raw_writer.append_aligned_row(row)
            self._last_std_frame = fA
            self._last_lab_frame = fB


class SensorManager:
    """传感器管理器"""
    def __init__(self, config, raw_writer: RawWriter, ring: RingBuffer):
        self.config = config
        self.sensors = {}
        self.threads = {}
        self.queues = {}
        self.running = False

        self.raw_writer = raw_writer
        self.ring = ring

        self.latest = {
            "liuzhou": None,
            "lab_six": None,
        }

        self._lock = threading.Lock()
        self.is_collecting = False
        self.periodic = None  # PeriodicAligner

    def initialize(self):
        """初始化传感器"""
        sensor_cfgs = self.config.get("sensor", {})

        # 标准六轴
        if "m8128b1" in sensor_cfgs:
            try:
                print("=== 初始化标准六轴力传感器 ===")
                sensor = create_std_sensor("m8128b1", sensor_cfgs["m8128b1"])
                if init_std_sensor(sensor):
                    self.sensors["liuzhou"] = sensor
                    self.queues["liuzhou"] = Queue()
                    print("✅ 标准六轴初始化成功")
                else:
                    print("❌ 标准六轴初始化失败")
            except Exception as e:
                print(f"❌ 标准六轴初始化异常: {e}")

        # 实验室六轴
        if "lab_six_axis" in sensor_cfgs:
            try:
                print("=== 初始化实验室六轴传感器 ===")
                sensor = create_lab_sensor( sensor_cfgs["lab_six_axis"])
                if init_lab_sensor(sensor):
                    self.sensors["lab_six"] = sensor
                    self.queues["lab_six"] = Queue()
                    print("✅ 实验室六轴初始化成功")
                else:
                    print("❌ 实验室六轴初始化失败")
            except Exception as e:
                print(f"❌ 实验室六轴初始化异常: {e}")

        return len(self.sensors) > 0

    def start(self, log_fn=None):
        """启动数据流与采集线程"""
        self.running = True
        for name, sensor in self.sensors.items():
            ok = False
            try:
                ok = sensor.start_stream()
            except Exception as e:
                if log_fn: 
                    log_fn.emit(f"❌ {name} start_stream 失败: {e}")
            if not ok:
                if log_fn: 
                    log_fn.emit(f"❌ {name} 启动数据流失败")
                continue

            t = threading.Thread(target=self._reader_thread,
                                 args=(name, sensor, self.queues[name], log_fn),
                                 daemon=True)
            t.start()
            self.threads[name] = t
            if log_fn: 
                log_fn.emit(f"✅ {name} 数据线程启动")

        # 启动对齐线程（仅当两类传感器都存在时）
        if ("liuzhou" in self.sensors) and ("lab_six" in self.sensors):
            align_cfg = self.config.get("align", {})
            period_ms = int(align_cfg.get("period_ms", 10))
            maxlen = int(align_cfg.get("maxlen", 200))
            min_ready = int(align_cfg.get("min_ready", 200))
            require_both_changed = bool(align_cfg.get("require_both_changed", True))
            self.periodic = PeriodicAligner(
                raw_writer=self.raw_writer,
                period_ms=period_ms,
                maxlen=maxlen,
                min_ready=min_ready,
                is_collecting_fn=lambda: self.is_collecting,
                require_both_changed=require_both_changed,
                debug_print=False
            )
            self.periodic.start()
            if log_fn:
                log_fn.emit(f"✅ 对齐线程启动 period={period_ms}ms maxlen={maxlen} min_ready={min_ready} bothChanged={require_both_changed}")

    def _reader_thread(self, name, sensor, q: Queue, log_fn):
        """数据读取线程"""
        frame_counter = 0
        while self.running:
            try:
                data_list = sensor.read_data()
                for dev_frame_no, groups in data_list:
                    for idx, data_tuple in enumerate(groups):
                        now_iso = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                        now_epoch = time.time()
                        now_mono = time.monotonic()

                        # 如果是标准六轴且只有6个通道，则为其补上 temp=0
                        final_data_tuple = data_tuple
                        if name == "liuzhou" and len(data_tuple) == 6:
                            final_data_tuple = data_tuple + (0,) # 补上温度值

                        q.put((now_iso.split(" ")[1], frame_counter, final_data_tuple))

                        with self._lock:
                            self.latest[name] = (now_iso.split(" ")[1], frame_counter, final_data_tuple)

                        if self.is_collecting:
                            raw_row = {
                                "ts_iso": now_iso,
                                "ts_epoch": f"{now_epoch:.6f}",
                                "frame": frame_counter,
                            }
                            # 两个传感器都使用相同的key
                            keys = ["fx", "fy", "fz", "mx", "my", "mz", "temp"]
                            for i, v in enumerate(final_data_tuple):
                                if i < len(keys):
                                    raw_row[keys[i]] = v
                            self.raw_writer.append_raw_row(name, raw_row)

                        self.ring.push(name, now_mono, frame_counter, final_data_tuple)
                        # 推送到对齐缓冲
                        if self.periodic is not None:
                            self.periodic.add(name, now_mono, frame_counter, final_data_tuple)
                        frame_counter += 1

                time.sleep(0.001)
            except Exception as e:
                if self.running and log_fn:
                    log_fn.emit(f"❌ {name} 读数据异常: {e}")
                time.sleep(0.1)

    def zero_both(self, log_fn=None, order=("liuzhou","lab_six"), post_delay_s=0.2):
        """同时清零"""
        start_ts = time.time()
        if log_fn: 
            log_fn.emit(f"=== 清零开始（顺序：标准六轴 → 实验室六轴）=== ")

        for name in order:
            sensor = self.sensors.get(name)
            if sensor is None:
                if log_fn: 
                    log_fn.emit(f"跳过 {name}（未连接）")
                continue
            try:
                if log_fn: 
                    log_fn.emit(f"→ 清零 {name} ...")
                ok = sensor.zero_channels()
                if ok:
                    if log_fn: 
                        log_fn.emit(f"✅ {name} 清零完成")
                else:
                    if log_fn: 
                        log_fn.emit(f"❌ {name} 清零返回失败")
            except Exception as e:
                if log_fn: 
                    log_fn.emit(f"❌ {name} 清零异常: {e}")

            time.sleep(post_delay_s)

        dur = time.time() - start_ts
        if log_fn: 
            log_fn.emit(f"=== 清零结束（总耗时 {dur:.2f}s）=== ")

    def stop(self, log_fn=None):
        """停止与收尾"""
        self.running = False
        # 先停止对齐线程
        try:
            if self.periodic is not None:
                self.periodic.stop()
        except Exception:
            pass
        for name, t in self.threads.items():
            try:
                t.join(timeout=0.5)
            except Exception:
                pass

        for name, sensor in self.sensors.items():
            try:
                sensor.stop_stream()
                sensor.disconnect()
                if log_fn: 
                    log_fn.emit(f"已关闭 {name}")
            except Exception as e:
                if log_fn: 
                    log_fn.emit(f"关闭 {name} 异常: {e}")




class PlotWindow(QMainWindow):
    """实时曲线显示窗口"""
    def __init__(self, sensor_manager, checkboxes, colors):
        super().__init__()
        self.sensor_manager = sensor_manager
        self.checkboxes = checkboxes
        self.colors = colors
        self.setWindowTitle("实时曲线显示")
        self.resize(800, 600)

        # 创建 PlotWidget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        self.plot_widget = pg.PlotWidget()
        self.plot_widget.showGrid(x=True, y=True)
        self.plot_widget.setLabel("left", "值")
        self.plot_widget.setLabel("bottom", "时间")
        layout.addWidget(self.plot_widget)

        # 初始化曲线
        self.curves = {name: self.plot_widget.plot(pen=pg.mkPen(color)) 
                      for name, color in self.colors.items()}
        self.plot_data = {name: [] for name in self.colors.keys()}
        
        # 添加时间轴数据
        self.time_data = {name: [] for name in self.colors.keys()}
        self.start_time = time.time()  # 记录开始时间
        
        # 设置最大数据点数
        self.max_points = 200

        # 定时器
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_plot)
        self.timer.start(100)

    def update_plot(self):
        """实时更新曲线"""
        current_time = time.time() - self.start_time
        
        with self.sensor_manager._lock:
            std_six_data = self.sensor_manager.latest["liuzhou"]
            lab_six_data = self.sensor_manager.latest["lab_six"]

        # 更新标准六轴数据
        if std_six_data:
            keys = ["std_fx", "std_fy", "std_fz", "std_mx", "std_my", "std_mz", "std_temp"]
            for i, key in enumerate(keys):
                if self.checkboxes[key].isChecked():
                    if i < len(std_six_data[2]):
                        self.plot_data[key].append(std_six_data[2][i])
                        self.time_data[key].append(current_time)
                        
                        if len(self.plot_data[key]) > self.max_points:
                            self.plot_data[key].pop(0)
                            self.time_data[key].pop(0)
                        
                        if len(self.time_data[key]) == len(self.plot_data[key]):
                            self.curves[key].setData(self.time_data[key], self.plot_data[key])
                else:
                    self.curves[key].setData([], [])

        # 更新实验室六轴数据
        if lab_six_data:
            keys = ["lab_fx", "lab_fy", "lab_fz", "lab_mx", "lab_my", "lab_mz", "lab_temp"]
            for i, key in enumerate(keys):
                if self.checkboxes[key].isChecked():
                    if i < len(lab_six_data[2]):
                        self.plot_data[key].append(lab_six_data[2][i])
                        self.time_data[key].append(current_time)
                        
                        if len(self.plot_data[key]) > self.max_points:
                            self.plot_data[key].pop(0)
                            self.time_data[key].pop(0)
                        
                        if len(self.time_data[key]) == len(self.plot_data[key]):
                            self.curves[key].setData(self.time_data[key], self.plot_data[key])
                else:
                    self.curves[key].setData([], [])


class MainWindow(QMainWindow):
    """主窗口"""
    log_signal = pyqtSignal(str)

    def __init__(self, manager: SensorManager, xlsx_path: str, raw_writer: RawWriter, ring: RingBuffer):
        super().__init__()
        self.mgr = manager
        self.xlsx_path = xlsx_path
        self.raw_writer = raw_writer
        self.ring = ring
        self.plot_window = None

        self.init_ui()
        self.setup_timers()
        
        # 连接日志信号
        self.log_signal.connect(self.log_print)

    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("双六轴传感器监视与对比")
        self.resize(980, 640)

        # 主布局
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 顶部状态栏
        self.status_label = QLabel("准备中…")
        self.status_label.setStyleSheet("QLabel { padding: 8px; }")
        main_layout.addWidget(self.status_label)

        # Excel文件信息
        excel_label = QLabel(f"Excel文件：{os.path.basename(self.xlsx_path)}")
        excel_label.setAlignment(Qt.AlignRight)
        excel_label.setStyleSheet("QLabel { padding: 8px; }")
        main_layout.addWidget(excel_label)

        # 按钮区域
        self.create_buttons(main_layout)

        # 复选框区域
        self.create_checkboxes(main_layout)

        # 传感器数据显示区域
        self.create_sensor_displays(main_layout)

        # 提示
        hint = QLabel("提示：点击『标定采样』 → 截取最近 1.0s 数据并写入 Sheet2_Aligned（均值）。")
        hint.setStyleSheet("QLabel { color: #555; padding: 8px; }")
        main_layout.addWidget(hint)

        # 日志区域
        self.create_log_area(main_layout)

    def create_buttons(self, main_layout):
        """创建按钮区域"""
        button_frame = QFrame()
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)

        # 采集按钮
        self.collect_btn = QPushButton("采集")
        self.collect_btn.clicked.connect(self.on_collect_clicked)
        button_layout.addWidget(self.collect_btn)

        # 保存按钮
        self.save_btn = QPushButton("保存")
        self.save_btn.setEnabled(False)
        self.save_btn.clicked.connect(self.on_save_clicked)
        button_layout.addWidget(self.save_btn)

        # 标定采样按钮
        self.calibrate_btn = QPushButton("标定采样")
        self.calibrate_btn.setEnabled(False)
        self.calibrate_btn.clicked.connect(self.on_snapshot)
        button_layout.addWidget(self.calibrate_btn)

        # 清零按钮
        self.zero_btn = QPushButton("清零")
        self.zero_btn.clicked.connect(self.on_zero_clicked)
        button_layout.addWidget(self.zero_btn)

        # 显示实时曲线按钮
        self.plot_btn = QPushButton("显示实时曲线")
        self.plot_btn.clicked.connect(self.show_plot_window)
        button_layout.addWidget(self.plot_btn)

        # 导出按钮
        self.export_btn = QPushButton("导出原始到Excel")
        self.export_btn.clicked.connect(self.on_export_raw)
        button_layout.addWidget(self.export_btn)

        # 退出按钮
        self.quit_btn = QPushButton("退出")
        self.quit_btn.clicked.connect(self.close)
        button_layout.addWidget(self.quit_btn)

    def create_checkboxes(self, main_layout):
        """创建复选框区域"""
        """创建复选框区域"""
        self.checkboxes = {}
        self.colors = {
            # 标准六轴颜色 (亮色系)
            "std_fx": "red", 
            "std_fy": "green", 
            "std_fz": "blue",
            "std_mx": "cyan", 
            "std_my": "magenta", 
            "std_mz": "yellow", 
            "std_temp": "white",
            # 实验室六轴颜色 (另一套区分色系)
            "lab_fx": "#FFA500",  # 橙色
            "lab_fy": "#800080",  # 紫色
            "lab_fz": "#A52A2A",  # 棕色
            "lab_mx": "#FFC0CB",  # 粉色
            "lab_my": "#808080",  # 灰色
            "lab_mz": "#00FFFF",  # 青色 (Aqua)
            "lab_temp": "#ADD8E6"   # 亮蓝色
        }

        checkbox_group = QGroupBox("曲线选择")
        # 主布局使用垂直布局，来容纳两个水平行
        main_checkbox_layout = QVBoxLayout(checkbox_group)
        main_layout.addWidget(checkbox_group)

        # 为标准六轴和实验室六轴分别创建水平布局
        std_checkbox_layout = QHBoxLayout()
        lab_checkbox_layout = QHBoxLayout()

        # 将两个水平布局添加到主垂直布局中
        main_checkbox_layout.addLayout(std_checkbox_layout)
        main_checkbox_layout.addLayout(lab_checkbox_layout)

        # 分别为两个传感器添加复选框
        for name, color in self.colors.items():
            checkbox = QCheckBox(f"{name} ({color})")
            checkbox.setChecked(False)
            self.checkboxes[name] = checkbox
            
            if name.startswith("std_"):
                std_checkbox_layout.addWidget(checkbox)
            else:
                lab_checkbox_layout.addWidget(checkbox)
        
        # 在每行末尾添加伸缩，使复选框靠左对齐
        std_checkbox_layout.addStretch(1)
        lab_checkbox_layout.addStretch(1)

    def create_sensor_displays(self, main_layout):
        """创建传感器数据显示区域"""
        panel_frame = QFrame()
        panel_layout = QVBoxLayout(panel_frame)
        main_layout.addWidget(panel_frame)

        # 标准六轴显示
        std_six_group = QGroupBox("标准六轴最新值 (Fx Fy Fz Mx My Mz Temp)")
        std_six_layout = QVBoxLayout(std_six_group)
        self.std_six_label = QLabel("—")
        self.std_six_label.setFont(QFont("Consolas", 11))
        std_six_layout.addWidget(self.std_six_label)
        panel_layout.addWidget(std_six_group)

        # 实验室六轴显示
        lab_six_group = QGroupBox("实验室六轴最新值 (Fx Fy Fz Mx My Mz Temp)")
        lab_six_layout = QVBoxLayout(lab_six_group)
        self.lab_six_label = QLabel("—")
        self.lab_six_label.setFont(QFont("Consolas", 11))
        lab_six_layout.addWidget(self.lab_six_label)
        panel_layout.addWidget(lab_six_group)

    def create_log_area(self, main_layout):
        """创建日志区域"""
        log_group = QGroupBox("日志")
        log_layout = QVBoxLayout(log_group)
        main_layout.addWidget(log_group)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 10))
        log_layout.addWidget(self.log_text)

    def setup_timers(self):
        """设置定时器"""
        # 队列轮询定时器
        self.queue_timer = QTimer()
        self.queue_timer.timeout.connect(self.poll_queues)
        self.queue_timer.start(50)

        # 界面刷新定时器
        self.refresh_timer = QTimer()
        self.refresh_timer.timeout.connect(self.refresh_latest)
        self.refresh_timer.start(150)

    def log_print(self, text: str):
        """日志打印函数"""
        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        self.log_text.append(f"[{ts}] {text}")

    def on_collect_clicked(self):
        """采集按钮点击"""
        self.mgr.is_collecting = True
        self.collect_btn.setEnabled(False)
        self.save_btn.setEnabled(True)
        self.calibrate_btn.setEnabled(True)
        self.log_print("开始采集数据...")

    def on_save_clicked(self):
        """保存按钮点击"""
        self.mgr.is_collecting = False
        self.collect_btn.setEnabled(True)
        self.save_btn.setEnabled(False)
        self.calibrate_btn.setEnabled(False)
        self.log_print("数据采集已停止。")




    def on_snapshot(self):
        """标定采样按钮点击：点击后开始采集接下来的 5s 数据"""
        if not self.mgr.is_collecting:
            self.log_print("⚠️ 请先点击『采集』按钮开始采集数据！")
            return

        # 防重入
        if hasattr(self, "_snapshot_running") and self._snapshot_running:
            self.log_print("⏳ 正在采集上一轮 5s 数据，请稍后再试…")
            return

        self._snapshot_running = True
        self._snapshot_window_s = 5.0
        self._snapshot_start_mono = RingBuffer.now_mono()
        self.log_print(f"⏳ 已开始采集接下来的 {self._snapshot_window_s:.1f}s 数据…")

        # 在这里安排 5s 后调用收尾函数
        QTimer.singleShot(int(self._snapshot_window_s * 1000), self._finalize_snapshot)
    
    def _finalize_snapshot(self):
        """结束 5s 采集窗口，汇总均值并写入明细"""
        try:
            window_s = getattr(self, "_snapshot_window_s", 5.0)
            t_beg = getattr(self, "_snapshot_start_mono", RingBuffer.now_mono())
            t_end = t_beg + window_s

            a = self.ring.slice("liuzhou", t_beg, t_end)
            b = self.ring.slice("lab_six",  t_beg, t_end)

            a_frames = len(a)
            b_frames = len(b)
            if a_frames == 0 and b_frames == 0:
                self.log_print("⚠️ 该 5s 窗口内无数据，不写入")
                return

            def mean_tuple(samples: List[Sample]) -> List[float]:
                if not samples:
                    return []
                n_ch = len(samples[0].data)
                acc = [0.0] * n_ch
                for s in samples:
                    for i, v in enumerate(s.data):
                        if i < n_ch:
                            acc[i] += float(v)
                return [x / len(samples) for x in acc]

            a_mean = mean_tuple(a)
            b_mean = mean_tuple(b)

            gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            mean_header = (
                ["snapshot_id", "gen_time", "win_s", "std_frames", "lab_frames"]
                + ["std_fx", "std_fy", "std_fz", "std_mx", "std_my", "std_mz", "std_temp"]
                + ["lab_fx", "lab_fy", "lab_fz", "lab_mx", "lab_my", "lab_mz", "lab_temp"]
            )
            
            all_data_header = (
                ["snapshot_id", "device", "ts_mono", "frame"]
                + ["fx", "fy", "fz", "mx", "my", "mz", "temp"]
            )

            wb = load_workbook(self.xlsx_path)

            # 均值表
            if "Sheet2_biaoding" in wb.sheetnames:
                ws_mean = wb["Sheet2_biaoding"]
                need_header_mean = (ws_mean.max_row <= 1 and (ws_mean["A1"].value is None))
            else:
                ws_mean = wb.create_sheet("Sheet2_biaoding")
                need_header_mean = True
            if need_header_mean:
                ws_mean.append(mean_header)

            snapshot_id = ws_mean.max_row - 1 if ws_mean.max_row > 1 else 0

            # 明细表
            if "Sheet2_all_data" in wb.sheetnames:
                ws_all = wb["Sheet2_all_data"]
                need_header_all = (ws_all.max_row <= 1 and (ws_all["A1"].value is None))
            else:
                ws_all = wb.create_sheet("Sheet2_all_data")
                need_header_all = True
            if need_header_all:
                ws_all.append(all_data_header)

            # 写入均值
            mean_values = [snapshot_id + 1, gen_time, window_s, a_frames, b_frames] + a_mean + b_mean
            ws_mean.append(mean_values)

            # 写入明细
            for s in a:
                vals = list(s.data[:7]) if len(s.data) >= 7 else [""] * 7
                ws_all.append([snapshot_id + 1, "liuzhou", f"{s.ts_mono:.6f}", s.frame] + vals)
            for s in b:
                vals = list(s.data[:7]) if len(s.data) >= 7 else [""] * 7
                ws_all.append([snapshot_id + 1, "lab_six", f"{s.ts_mono:.6f}", s.frame] + vals)

            wb.save(self.xlsx_path)
            self.log_print(f"✅ 5s 采集完成：#{snapshot_id + 1} 标准六轴={a_frames} 实验室六轴={b_frames}")
        except Exception as e:
            self.log_print(f"❌ 5s 采集/写入失败：{e}")
        finally:
            self._snapshot_running = False



    def on_zero_clicked(self):
        """清零按钮点击"""
        self.log_print("开始清零所有传感器...")
        self.mgr.zero_both(log_fn=self.log_signal)
        self.log_print("清零完成。")

    def on_export_raw(self):
        """导出原始数据按钮点击"""
        self.log_print("开始导出原始数据到 Excel...")
        try:
            self.raw_writer.export_csv_to_sheets()
            self.log_print(f"✅ 原始数据已成功导出到 {self.xlsx_path}")
        except Exception as e:
            self.log_print(f"❌ 导出原始数据失败: {e}")

    def show_plot_window(self):
        """显示实时曲线窗口"""
        if self.plot_window is None:
            self.plot_window = PlotWindow(self.mgr, self.checkboxes, self.colors)
        self.plot_window.show()

    def refresh_latest(self):
        """刷新最新传感器数据"""
        with self.mgr._lock:
            # 标准六轴
            if self.mgr.latest["liuzhou"]:
                ts, frame, data = self.mgr.latest["liuzhou"]
                # 数据在读取时已经补全为7通道
                if len(data) >= 7:
                    Fx, Fy, Fz, Mx, My, Mz, Temp = data[:7]
                    self.std_six_label.setText(
                        f"#{frame:05d} Fx={Fx:8.6f} Fy={Fy:8.6f} Fz={Fz:8.6f} "
                        f"Mx={Mx:8.6f} My={My:8.6f} Mz={Mz:8.6f} Temp={Temp:6.1f}"
                    )
            else:
                self.std_six_label.setText("—")

            # 实验室六轴
            if self.mgr.latest["lab_six"]:
                ts, frame, data = self.mgr.latest["lab_six"]
                if len(data) >= 7:
                    Fx, Fy, Fz, Mx, My, Mz, Temp = data[:7]
                    self.lab_six_label.setText(
                        f"#{frame:05d} Fx={Fx:8.6f} Fy={Fy:8.6f} Fz={Fz:8.6f} "
                        f"Mx={Mx:8.6f} My={My:8.6f} Mz={Mz:8.6f} Temp={Temp:6.3f}"
                    )
            else:
                self.lab_six_label.setText("—")



    def poll_queues(self):
        """轮询传感器队列"""
        for name, q in self.mgr.queues.items():
            took = 0
            while took < 20:
                try:
                    ts_str, frame_no, data = q.get_nowait()
                except Empty:
                    break
                
                if not self.mgr.is_collecting:
                    continue
                
                # 数据在读取时已经补全为7通道
                if name in ["liuzhou", "lab_six"] and len(data) >= 7:
                    Fx, Fy, Fz, Mx, My, Mz, Temp = data[:7]
                    line = (f"[{name}] {ts_str} #{frame_no:05d} "
                        f"Fx={Fx:8.4f} Fy={Fy:8.4f} Fz={Fz:8.4f} "
                        f"Mx={Mx:8.4f} My={My:8.4f} Mz={Mz:8.4f} Temp={Temp:6.1f}")
                else:
                    line = f"[{name}] {ts_str} #{frame_no:05d} {data}"
                # self.log_print(line) # 默认关闭日志刷屏
                took += 1

    def closeEvent(self, event):
        """窗口关闭事件"""
        try:
            self.mgr.stop()
        except Exception:
            pass
        try:
            self.raw_writer.close()
        except Exception:
            pass
        print("[CLOSE] 设备已关闭")
        event.accept()


def main():
    """主函数"""
    print("=== PyQt 版双六轴传感器监视与对比 ===")

    app = QApplication(sys.argv)

    # 加载配置
    cfg = load_config()
    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = os.path.join(os.getcwd(), f"data_{run_ts}.xlsx")

    raw_writer = RawWriter(xlsx_path, flush_every_n=200)
    ring = RingBuffer(max_seconds=120.0)

    mgr = SensorManager(cfg, raw_writer=raw_writer, ring=ring)
    if not mgr.initialize():
        print("❌ 没有传感器初始化成功，程序退出。")
        try:
            raw_writer.close()
        except Exception:
            pass
        return

    # 创建主窗口
    main_window = MainWindow(mgr, xlsx_path=xlsx_path, raw_writer=raw_writer, ring=ring)
    main_window.status_label.setText("初始化完成，正在启动数据流…")
    mgr.start(log_fn=main_window.log_signal)
    main_window.status_label.setText("运行中：数据持续接收。")
    main_window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()